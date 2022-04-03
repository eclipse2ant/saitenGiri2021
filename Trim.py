import csv
import os
import sys
import shutil
import glob
import pathlib
import openpyxl
import imghdr
from turtle import pos
from PIL import Image, ImageTk, ImageDraw, ImageFont  # 外部ライブラリ
from tkinter import messagebox


import mFileUtil as fu

datafile = "./setting/trimData.csv"
#output_dir = "./setting/output"
# トリミング前の画像の格納先
ORIGINAL_FILE_DIR = "./setting/input"
# トリミング後の画像の格納先
TRIMMED_FILE_DIR = "./setting/output"

extlist =  ['jpg', "jpeg", "png", "PNG", "JPEG", "JPG", "gif"]

def make_trim_data(file, dir, data):
    return {'file': file, 'dir': dir, 'data': data}

def do_trim(image, output_dir, pos, file): 
    title , left , top , right , bottom = pos
    # 出力フォルダのパス
    # もしトリミング後の画像の格納先が存在しなければ作る
    fu.if_mkdir(fu.addpath(output_dir, title))
    im_trimmed = image.crop((int(left), int(top), int(right), int(bottom)))
    # qualityは95より大きい値は推奨されていないらしい
    im_trimmed.save(fu.addpath(fu.addpath(output_dir, title), file), quality=95)
    print(fu.addpath(fu.addpath(output_dir, title), file))
    print("___"+ title + "を斬り取りました。" )
    print("********************************")
    

def trim(t_data):
  
    # トリミングされたimageオブジェクトを取得
    image = Image.open(fu.addpath(t_data['dir'],t_data['file']))
    print(t_data['file'] + "を斬ります" )
    
    for pos in t_data['data']:
        do_trim(image, TRIMMED_FILE_DIR, pos, t_data['file'])

def allTrim():
    data = fu.readCSV(datafile)
    try:
        shutil.rmtree(TRIMMED_FILE_DIR)
    except OSError as err:
        pass
    
    if data == 0:
        print('終了', 'どうやって斬ればいいかわかりません。\nまずはどこを斬るかを決めてください。')
    #	messagebox.showinfo('終了', 'どうやって斬ればいいかわかりません。\nまずはどこを斬るかを決めてください。')
        return 0

  # 画像ファイル名を取得	
  # 特定の拡張子のファイルだけを採用。実際に加工するファイルの拡張子に合わせる
    files = fu.ext_filter(os.listdir(ORIGINAL_FILE_DIR), extlist)
    #print(files)
    try:
        for file in files:
            trim(make_trim_data(file,  ORIGINAL_FILE_DIR, data))
    except:
            print(
        'エラー', 'エラーが検出されました。中断します。\n\n' + str(sys.stderr))

#			messagebox.showinfo(
#        'エラー', 'エラーが検出されました。中断します。\n\n' + str(sys.stderr))
            try:
                shutil.rmtree("./setting/output")
            except OSError as err:
                pass
            return 0

    # nameフォルダの中身をリサイズ
    # maxheight以上のときは、小さくする。
    maxheight = 50
    files = glob.glob("./setting/output/name/*")
    img = Image.open(files[0])
    namew, nameh = img.size
    if nameh > maxheight:
        rr = nameh / maxheight
        for f in files:
            img = Image.open(f)
            img = img.resize((int(namew / rr), int(nameh/rr)))
            img.save(f)
    output_name_sh()
    messagebox.showinfo('斬りました', '全員分の解答用紙を斬りました。')

def output_name_sh():
    
    # 定数設定
    SHEET_TITLE = '採点シート'  # シート名の設定
    RESULT_FILE_NAME = './setting/saiten.xlsx'  # 結果を保存するファイル名

    # 変数
    max_height = []  # 各行の画像の高さの最大値を保持

    def get_file_names(set_dir_name):
        """
        ディレクトリ内のファイル名取得（ファイル名のみの一覧を取得）
        """
        file_names = os.listdir(set_dir_name)
        temp_full_file_names = [os.path.join(set_dir_name, file_name) for file_name in file_names if os.path.isfile(
            os.path.join(set_dir_name, file_name))]  # ファイルかどうかを判定
        return temp_full_file_names

    def attach_img(target_full_file_names, set_column_idx, set_dir_name):
        """
        画像を呼び出して、Excelに貼り付け
        """
        set_row_idx = 1
        column_letter = "B"
        # 各列の1行目に、貼り付ける画像があるディレクトリ名を入力
        ws.cell(row=1, column=set_column_idx).value = "画像"
        ws.cell(row=1, column=1).value = "ファイル名"  # ファイル名
        ws.cell(row=1, column=3).value = "生徒番号"  # 出席番号
        ws.cell(row=1, column=4).value = "名前"  # 名前
        max_width = 0  # 画像の幅の最大値を保持するための変数
        target_full_file_names.sort()  # ファイル名でソート
        for target_file in target_full_file_names:
            p = pathlib.Path(target_file)
            target_file = p.resolve()
            if imghdr.what(target_file) != None:  # 画像ファイルかどうかの判定
                img = openpyxl.drawing.image.Image(target_file)
                #print('[' + column_letter + '][' + str(set_row_idx+1) + ']' + target_file + 'を貼り付け')

                # 画像のサイズを取得して、セルの大きさを合わせる（画像同士が重ならないようにするため）
                size_img = Image.open(target_file)
                width, height = size_img.size
                if max_width < width:
                    max_width = width
                # 配列「max_height」において、「set_row_idx」番目の要素が存在しなければ、挿入
                if not max_height[set_row_idx-1:set_row_idx]:
                    max_height.insert(set_row_idx-1, height)
                if max_height[set_row_idx-1] < height:
                    max_height[set_row_idx-1] = height
                ws.row_dimensions[set_row_idx +
                                  1].height = max_height[set_row_idx-1] * 0.75
                ws.column_dimensions[column_letter].width = int(
                    max_width) * 0.13

                # セルの行列番号から、そのセルの番地を取得
                cell_address = ws.cell(
                    row=set_row_idx + 1, column=set_column_idx).coordinate
                img.anchor = cell_address
                ws.add_image(img)  # シートに画像貼り付け
                ws.cell(row=set_row_idx + 1,
                        column=1).value = os.path.basename(target_file)

            set_row_idx += 1

    # ワークブック設定
    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]  # 1番目のシートを編集対象にする
    ws.title = SHEET_TITLE  # 1番目のシートに名前を設定

    # 貼り付ける画像を置いておくルートディレクトリ内のディレクトリ名を再帰的に取得
    dir_name = "./setting/output/name"

    column_idx = 2

    f_names = get_file_names(dir_name)  # ファイル名取得
    attach_img(f_names, column_idx, dir_name)  # 画像貼り付け設定

    # ファイルへの書き込み
    wb.save(RESULT_FILE_NAME)



#allTrim()



      