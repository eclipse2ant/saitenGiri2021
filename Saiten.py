from asyncio.windows_events import NULL
from tkinter import *
from tkinter import messagebox
from PIL import Image, ImageTk, ImageDraw, ImageFont  # 外部ライブラリ

import os
import csv
import shutil
import openpyxl

from attr import attr
import MyFileUtil as fu
from Windows import Windows

class Saiten(Windows):
    
    OUTPUT_PATH = "./setting/output/"

    
    def init(self):     
        self.lb = Listbox(self.tk, selectmode='single', height=20, width=20)
        clcounter = 0
        for dir in self.get_dirs(self.OUTPUT_PATH):
            if not dir == "name":
                misaiten = len([f for f in next(os.walk("./setting/output/" + dir))[2] if not f.startswith('.')])
                self.lb.insert(END, dir)
                if misaiten == self.maxNinzu():
                    self.lb.itemconfig(clcounter, {'bg': 'white'})
                elif misaiten == 0:
                    self.lb.itemconfig(clcounter,  {'bg': 'gray'})
                else:
                    self.lb.itemconfig(clcounter,  {'bg': 'pale green'})
                clcounter = clcounter + 1

        self.lb.grid(row=0, column=0)
        # Scrollbar
        scrollbar = Scrollbar(
            self.tk,
            orient=VERTICAL,
            command=self.lb.yview)
        self.lb['yscrollcommand'] = scrollbar.set
        scrollbar.grid(row=0, column=1,  sticky=(N, S, W))

        button_frame = Frame(self.tk)
        button_frame.grid(row=0, column=1, sticky=W +
                      E + N + S, padx=30, pady=30)

        siroKaisetsu = Label(button_frame, text="未採点", bg="white").pack(
            side=TOP, fill=X)
        midoriKaisetsu = Label(button_frame, text="採点中", bg="pale green").pack(
            side=TOP, fill=X)
        grayKaisetsu = Label(button_frame, text="採点終了", bg="gray").pack(
            side=TOP, fill=X)

        button1 = Button(
            button_frame, text='採点する', width=15, height=3,
            command=lambda: self.show_selection()).pack(expand=True)

        totopB = Button(
            button_frame, text='Topに戻る', width=15, height=3,
            command=self.backTop).pack()
        
    # macにおける、.DS_storeを無視してカウントする。
    def maxNinzu(self):
        return len([f for f in next(os.walk("./setting/input/"))[2] if not f.startswith('.')])

    def show_selection(self):
        for i in self.lb.curselection():
            print(self.lb.get(i))
            attr = {'title': "採点中...", 'geometry': "1000x800", 'bg': "grey90",
                    'Qnum': str(self.lb.get(i))
                    }
            print(attr)
            SiwakeApp(attr).do()
            
    def backTop(self):
        self.tk.destroy()
        
    def get_files(self, path):
        return fu.get_files(path)
    
    # outputの中のフォルダを取得
    def get_dirs(self, path):
        return fu.get_dirs(path)
        
class SiwakeApp(Windows):
    def init(self):
        # グローバル変数
        #global filename_lst
        #global assort_file_list
        #global assort_dict

       # global image_canvas

        ''''
        global img_lst, tk_img_lst
        global filename_lst
        global assort_file_list
        global assort_dict

        global tex_var
        #global image_canvas
        # global assort_btn
        global assort_t_var

        global img_num
        global f_basename
        global siwake_win
        global saitenCount
        global f_basename
        '''
        
        self.assort_btn = NULL
        self.image_canvas = NULL
        self.filename_lst = NULL
        self.file_list = NULL
        self.assort_t_var = NULL
        self.img_num  = NULL
        self.saitenCount = NULL
        self.dir_name = NULL
        self.f_dir = NULL
        
        self.img_lst, self.tk_img_lst = [], []
        self.filename_lst = []
        self.assort_file_list = []
        self.assort_dict = {}
        self.cbln = []
        self.chk = []
        
        print("siwakeApp IN " + str(self.attr['Qnum']))

        self.img_num = 0
        self.f_basename = ""

        siwake_frame = Frame(self.tk)
        siwake_frame.grid(column=0, row=0)
        button_siwake_frame = Frame(self.tk)
        button_siwake_frame.grid(
            column=1, row=0, sticky=W + E + N + S)

        # キャンバス描画設定
        self.image_canvas = Canvas(siwake_frame,
                                  bg="green",
                                  width=640,
                                  height=480)

        self.image_canvas.pack(expand=True, fill="both")
        
        # 仕分け結果表示
        self.assort_t_var = StringVar(siwake_frame)
        self.assort_t_var.set("1 ~ 9のキーで点数を入力してください\n[space]で採点をskipします")
        assort_label = Label(
        siwake_frame, textvariable=self.assort_t_var, font=("Meiryo UI", 30), bg="white",  relief="sunken")
        assort_label.pack()
        
        # ファイル名ラベル描画設定
        self.tex_var = StringVar(siwake_frame)
        self.tex_var.set("ファイル名")

        lbl = Label(siwake_frame, textvariable=self.tex_var,
                                font=("Meiryo UI", 20))
        lbl["foreground"] = "gray"
        lbl.pack()
        
        # 右左キーで画像送りする動作設定
        self.tk.bind("<Key-Right>", self.next_img)
        self.tk.bind("<Key-Left>", self.prev_img)
        # 「Ctrl」+「P」で画像表示
        self.tk.bind("<Control-Key-p>", self.image_show)

        # 数字キーで仕分け対象設定
        self.tk.bind("<Key>", self.file_assort)
        
        # 仕分け実行ボタン
        self.assort_btn = Button(
            button_siwake_frame, text="採点実行",  height=3, width=15)
        self.assort_btn.bind("<Button-1>", self.assort_go)

        # ファイル名ラベル描画設定
        self.saitenCount = StringVar(button_siwake_frame)
        self.saitenCount.set("")
        ikutsuLb = Label(
            button_siwake_frame, textvariable=self.saitenCount, font=("Meiryo UI", 20))
        ikutsuLb.pack(side=TOP)

        exit_button = Button(
            button_siwake_frame, text="トップに戻る\n保存はされません", height=3, width=15,  command=self.exit_siwake)
        exit_button.pack()

        backfigB = Label(siwake_frame, text="←前へ\nキーボードの←ボタン", font=(
            "Meiryo UI", 20)).pack(side=LEFT, expand=TRUE)
        nextfigB = Label(siwake_frame, text="次へ→\nキーボードの→ボタン", font=(
        "Meiryo UI", 20)).pack(side=RIGHT, expand=TRUE)

        ## 禁則処理のゾーン-------------
        setumeiBun1 = Label(button_siwake_frame , text = "入力可能な点数にチェックをつけてください。" ).pack(side = TOP)
        setumeiBun2 = Label(button_siwake_frame , text = "誤った数字キーを押すのを防ぎます。").pack(side = TOP)
        chkfonts = ("Meiryo UI", 10)
        for i in range(0,10):
            self.cbln.append(BooleanVar(master = self.tk))
            self.chk.append(Checkbutton(master = button_siwake_frame,  variable=self.cbln[i] ,text=str(i) , font = chkfonts).pack(side = TOP))
        
        # 読み込みボタン描画設定
        self.load_file()
        
        #siwake_win.mainloop


    def file_assort(self, event):
        tokutenList = []
        for i in range(0,10):
            if self.cbln[i].get():
                tokutenList.append(str(i))
        print("入力可能な点数は" + str(tokutenList))

        if str(event.keysym) in tokutenList:
            self.assort_dict[self.filename_lst[self.img_num]] = str(event.keysym)
        elif str(event.keysym) == "space":
            self.assort_dict[self.filename_lst[self.img_num]] = str("skip")
        elif str(event.keysym) in  ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9"]:
            self.assort_dict[self.filename_lst[self.img_num]] = str("その点数は入力できません。\n右のチェックを確認してください。")
        else:
            self.assort_dict[self.filename_lst[self.img_num]] = str("そのキーは対応してません。")

        # ラベリングを表示
        if self.filename_lst[self.img_num] in self.assort_dict:
            self.assort_t_var.set(self.assort_dict[self.filename_lst[self.img_num]])
        else:
            self.assort_t_var.set("")

        print(self.assort_dict[self.filename_lst[self.img_num]])


    def saiten2xlsx(self):
        def readCSV():
            # もしcsvが無ければ、全部止める
            if os.path.isfile("./setting/trimData.csv") == False:
                return 0
            else:
                with open('./setting/trimData.csv') as f:
                    reader = csv.reader(f)
                    data = [row for row in reader]
                    data.pop(0)
                    return data

        def setTensu(figname, qname, tensu):

            qCol = int(qname[-4:]) + 3
            ws.cell(1, qCol + 1).value = qname

            MIN_COL = 1
            MIN_ROW = 2

            MAX_COL = 1
            MAX_ROW = ws.max_row

            # 範囲データを順次処理
            for row in ws.iter_rows(min_col=MIN_COL, min_row=MIN_ROW, max_col=MAX_COL, max_row=MAX_ROW):
                for cell in row:
                    try:
                        # 該当セルの値取得
                        cell_value = cell.value
                        if figname == cell_value:
                            o = cell.offset(0, qCol)
                            try:
                                o.value = int(tensu)
                            except:
                                o.value = tensu
                    except:
                        pass

        data = readCSV()

        xlPath = "./setting/saiten.xlsx"
        wb = openpyxl.load_workbook(xlPath)
        ws = wb["採点シート"]

        while data:
            title, left, top, right, bottom = data.pop(0)
            if title == "name":
                continue
            qpath = "./setting/output/" + title
            for curDir, dirs, files in os.walk(qpath):
                if files:
                    for f in files:
                        tensu = os.path.basename(os.path.dirname(curDir + "/" + f))
                        if not tensu == title:
                            setTensu(figname=f, qname=title, tensu=tensu)
                        else:
                            setTensu(figname=f, qname=title, tensu="未")
        wb.save(xlPath)


 
 # フォルダ分け実行 - - - - - - - - - - - - - - - - - - - - - - - -
    
    def assort_go(self, event):

        for f in self.assort_dict:
        # 仕分け前後のファイル名・フォルダ名を取得
        # assort_dict[f]が[0~1]なら、フォルダを作る。      
        
            print(self.assort_dict[f])
            if self.assort_dict[f] in ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9"]:
                self.f_dir = os.path.dirname(f)
                self.f_basename = os.path.basename(f)
                new_dir = os.path.join(self.f_dir, self.assort_dict[f])
                new_path = os.path.join(new_dir, self.f_basename)

                # ディレクトリの存在チェック
                if not os.path.exists(new_dir):
                    os.mkdir(new_dir)
                # ファイルの移動実行
                shutil.move(f, new_path)

                print(new_path)
            else:
                pass
        self.saiten2xlsx()
        messagebox.showinfo("採点保存", "ここまでの採点結果を保存しました。\nskipした項目は、採点されていません。")
        self.tk.destroy()
 
 
    def exit_siwake(self):
        ret = messagebox.askyesno('終了します', '採点を中断し、ホームに戻っても良いですか？')
        if ret == True:
            self.tk.destroy()


    def load_file(self):

        # ファイルを読み込み
        self.tex_var.set("ファイルを読み込んでいます...")
        self.dir_name = "./setting/output/" + self.attr['Qnum']
        if not self.dir_name == None:
            self.file_list = fu.folder_walker(self.dir_name)

        # ファイルから読み込める画像をリストに列挙
        for f in self.file_list:
            try:
                print("__読み込み中..." + str(f))
                self.img_lst.append(Image.open(f))
                self.filename_lst.append(f)
            except:
                pass

        if not self.img_lst:
            self.tex_var.set("読み込む画像がありません。\n採点は終了しています。")

        # ウィンドウサイズに合わせてキャンバスサイズを再定義
        # window_resize()

        # 画像変換
        for f in self.img_lst:

            # キャンバス内に収まるようリサイズ
            resized_img = self.img_resize_for_canvas(f, self.image_canvas, expand=True)

            # tkinterで表示できるように画像変換
            self.tk_img_lst.append(ImageTk.PhotoImage(
                image=resized_img, master=self.image_canvas))

        # キャンバスの中心を取得
        c_width_half = round(int(self.image_canvas["width"]) / 2)
        c_height_half = round(int(self.image_canvas["height"]) / 2)

        # キャンバスに表示
        self.img_num = 0
        self.item = self.image_canvas.create_image(
            c_width_half, c_height_half,  image=self.tk_img_lst[0], anchor=CENTER)
        # ラベルの書き換え
        self.tex_var.set(self.filename_lst[self.img_num])
        self.saitenCount.set(str(self.img_num+1) + "/" + str(len(self.filename_lst)))

        # 仕分け実行ボタンの配置
        self.assort_btn.pack(expand=True)


    def image_show(self, event):
            self.img_lst[self.img_num].show()



    def prev_img(self, event):

    # 画像が最初でないか判定
        if self.img_num <= 0:
            pass
        else:
            # 表示中の画像No.を更新して表示
            self.img_num -= 1
            self.image_canvas.itemconfig(self.item, image=self.tk_img_lst[self.img_num])
            # ラベルの書き換え
            self.tex_var.set(self.filename_lst[self.img_num])
            self.saitenCount.set(str(self.img_num+1) + "/" + str(len(self.filename_lst)))
        # ラベリングを表示
        if self.filename_lst[self.img_num] in self.assort_dict:
            self.assort_t_var.set(self.assort_dict[self.filename_lst[self.img_num]])
        else:
            self.assort_t_var.set("")

    
    def next_img(self, event):

        # 読み込んでいる画像の数を取得
        img_count = len(self.tk_img_lst)

        # 画像が最後でないか判定
        if self.img_num >= img_count - 1:
            pass
        else:
            # 表示中の画像No.を更新して表示
            self.img_num += 1
            self.image_canvas.itemconfig(self.item, image=self.tk_img_lst[self.img_num])
            # ラベルの書き換え
            self.tex_var.set(self.filename_lst[self.img_num])
            self.saitenCount.set(str(self.img_num+1) + "/" + str(len(self.filename_lst)))
            # ラベリングを表示
            if self.filename_lst[self.img_num] in self.assort_dict:
                self.assort_t_var.set(self.assort_dict[self.filename_lst[self.img_num]])
            else:
                self.assort_t_var.set("")

    # キャンバスサイズに合わせて画像を縮小 - - - - - - - - - - - - - - - - - - - -
    def img_resize_for_canvas(self, img, canvas, expand=False):

        size_retio_w = int(canvas["width"]) / img.width
        size_retio_h = int(canvas["height"]) / img.height

        if expand == True:
            size_retio = min(size_retio_w, size_retio_h)
        else:
            size_retio = min(size_retio_w, size_retio_h, 1)

        resized_img = img.resize((round(img.width * size_retio),
                              round(img.height * size_retio)))
        return resized_img



    def exit_siwake(self):
        ret = messagebox.askyesno('終了します', '採点を中断し、ホームに戻っても良いですか？')
        if ret == True:
            self.tk.destroy()

